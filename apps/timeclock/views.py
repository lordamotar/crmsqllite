from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.utils import timezone
from django.http import HttpResponse
import io
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from .models import WorkSession, WorkDayMark
from django.contrib.auth import get_user_model
from .permissions import CanViewTimeclockReports


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def start_work(request):
    user = request.user
    open_session = WorkSession.objects.filter(user=user, is_closed=False).first()
    if open_session:
        return Response({'status': 'already_started', 'session_id': open_session.id})
    session = WorkSession.objects.create(
        user=user,
        start_time=timezone.now(),
        last_activity=timezone.now(),
        created_via='manual',
    )
    return Response({'status': 'started', 'session_id': session.id, 'start_time': session.start_time})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def stop_work(request):
    user = request.user
    session = WorkSession.objects.filter(user=user, is_closed=False).first()
    if not session:
        return Response({'status': 'no_open_session'}, status=400)
    session.close(end_time=timezone.now())
    return Response({'status': 'stopped', 'session_id': session.id, 'end_time': session.end_time})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def heartbeat(request):
    user = request.user
    session = WorkSession.objects.filter(user=user, is_closed=False).first()
    if not session:
        # Не автосоздаём сессию, чтобы "Завершить работу" не возобновляло её
        return Response({'status': 'no_open_session'})
    session.last_activity = timezone.now()
    session.save(update_fields=['last_activity'])
    return Response({'status': 'ok', 'session_id': session.id, 'last_activity': session.last_activity})


@api_view(['GET'])
@permission_classes([IsAuthenticated, CanViewTimeclockReports])
def export_timeclock_xlsx(request):
    date_from = request.GET.get('from')
    date_to = request.GET.get('to')
    uid = request.GET.get('user_id')

    # Если даты не переданы — берём текущий месяц
    if not date_from or not date_to:
        now = timezone.localtime()
        month_start = now.replace(day=1).date()
        # до сегодняшнего дня включительно
        df, dt = month_start, now.date()
    else:
        from datetime import datetime
        try:
            def make_naive_date(s: str):
                # поддержка ISO-строк вида YYYY-MM-DDTHH:MM:SS+TZ
                return datetime.strptime(s[:10], '%Y-%m-%d').date()
            df = make_naive_date(date_from)
            dt = make_naive_date(date_to)
        except ValueError:
            return Response({'detail': 'Invalid date format, expected YYYY-MM-DD or ISO8601'}, status=400)
    
    try:
        def make_naive(dt):
            if not dt:
                return None
            # Преобразуем к локальному и убираем tzinfo для Excel
            local = timezone.localtime(dt, timezone.get_current_timezone())
            return local.replace(tzinfo=None)
    except Exception:
        return Response({'detail': 'Date handling error'}, status=400)

    qs = WorkSession.objects.filter(start_time__date__gte=df, start_time__date__lte=dt)
    if uid:
        qs = qs.filter(user__id=uid)

    # Фильтрация пользователей по отделу/филиалу
    current_user = request.user
    User = get_user_model()
    
    # Если не админ и не указан конкретный user_id - фильтруем по отделу
    if not uid and not current_user.is_superuser:
        # Если есть подчиненные (начальник отдела) - свои + подчиненные
        subordinates = current_user.get_subordinates()
        if subordinates.exists():
            allowed_user_ids = [current_user.id] + list(subordinates.values_list('id', flat=True))
            qs = qs.filter(user_id__in=allowed_user_ids)
        else:
            # Обычный менеджер - только свои сессии
            qs = qs.filter(user=current_user)

    # Попытка: использовать шаблон other/Табель.xlsx
    from django.conf import settings
    import os
    import logging

    logger = logging.getLogger(__name__)
    template_path = os.path.join(settings.BASE_DIR, 'other', 'Табель.xlsx')
    use_template = os.path.exists(template_path)
    
    logger.info(f'Template path: {template_path}, exists: {use_template}')

    try:
        if use_template:
            logger.info(f'Using template: {template_path}')
            wb = load_workbook(template_path)
            ws = wb.active
            logger.info(f'Template loaded, active sheet: {ws.title}')

            # Заполняем по маппингу: дни D:AH, итог AI, строки начиная с 6 — реальные пользователи
            if uid:
                users = User.objects.filter(id=uid)
            else:
                # Фильтруем пользователей: только те, кто есть в базе и в отделе
                # Сначала получаем ID пользователей из сессий
                user_ids_from_sessions = qs.values_list('user_id', flat=True).distinct()
                
                # Базовый запрос - только активные пользователи из базы
                users_qs = User.objects.filter(
                    id__in=user_ids_from_sessions,
                    is_active=True
                )
                
                # Дополнительная фильтрация по отделу если не админ
                if not current_user.is_superuser:
                    subordinates = current_user.get_subordinates()
                    if subordinates.exists():
                        allowed_user_ids = [current_user.id] + list(subordinates.values_list('id', flat=True))
                        users_qs = users_qs.filter(id__in=allowed_user_ids)
                    else:
                        users_qs = users_qs.filter(id=current_user.id)
                
                # Убеждаемся что пользователи действительно существуют в базе
                users = users_qs.order_by('last_name', 'first_name')

            start_row = 6
            max_clear_rows = 200
            start_col = column_index_from_string('D')
            end_col = column_index_from_string('AH')
            # Определяем месяц шаблона: используем месяц который содержит больше дней из периода
            from datetime import date
            import calendar as _cal
            
            # Если период в пределах одного месяца - используем этот месяц
            if df.year == dt.year and df.month == dt.month:
                target_year, target_month = df.year, df.month
            else:
                # Если период пересекает месяцы - используем месяц даты "to"
                target_year, target_month = dt.year, dt.month
            
            days_in_month = _cal.monthrange(target_year, target_month)[1]
            
            # Обновим заголовок месяца в шаблоне (обычно в строке 1-2)
            # Ищем ячейки с текстом "Октябрь", "Ноябрь" и т.д. и обновляем
            month_names = {
                1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
                5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
                9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
            }
            month_name = month_names.get(target_month, '')
            
            # Обновляем заголовок (обычно в A1 или A2)
            for row in [1, 2]:
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        # Заменяем старые названия месяцев на новое
                        for old_month in month_names.values():
                            if old_month in str(cell.value):
                                cell.value = str(cell.value).replace(
                                    old_month, month_name
                                ).replace(
                                    '2025', str(target_year)
                                )
            
            # Разъединяем объединенные ячейки в строках 4-5 (дни месяца) перед записью
            # Собираем все объединенные диапазоны, которые затрагивают строки 4-5 и колонки D:AH
            merged_to_unmerge = []
            for merged_range in list(ws.merged_cells.ranges):
                # Проверяем, затрагивает ли объединение строки 4-5 и колонки дней
                if (merged_range.min_row <= 5 and merged_range.max_row >= 4 and
                    merged_range.min_col <= end_col and merged_range.max_col >= start_col):
                    merged_to_unmerge.append(merged_range)
            
            # Разъединяем найденные диапазоны
            for merged_range in merged_to_unmerge:
                try:
                    ws.unmerge_cells(str(merged_range))
                except Exception as e:
                    logger.warning(f'Could not unmerge {merged_range}: {e}')
            
            # Обновляем номера дней (1-31) в строке 4 и дни недели в строке 5
            # Строка 3 оставляем пустой
            # Дни недели: пн=0, вт=1, ср=2, чт=3, пт=4, сб=5, вс=6
            weekday_names = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']
            
            # Обновляем номера дней и дни недели в колонках D:AH (дни 1-31)
            for day_num in range(1, days_in_month + 1):
                col = start_col + (day_num - 1)
                day_date = date(target_year, target_month, day_num)
                weekday_name = weekday_names[day_date.weekday()]
                
                # Обновляем номер дня в строке 4 (числа 1-31)
                try:
                    ws.cell(row=4, column=col, value=day_num)
                except Exception as e:
                    logger.warning(f'Could not write to row 4, col {col}: {e}')
                
                # Обновляем день недели в строке 5
                try:
                    ws.cell(row=5, column=col, value=weekday_name)
                except Exception as e:
                    logger.warning(f'Could not write to row 5, col {col}: {e}')

            # Очистим диапазон строк (ФИО, Должность, дни, итог)
            for r in range(start_row, start_row + max_clear_rows):
                ws.cell(row=r, column=column_index_from_string('B'), value=None)
                ws.cell(row=r, column=column_index_from_string('C'), value=None)
                for c in range(start_col, end_col + 1):
                    ws.cell(row=r, column=c).value = None
                ws.cell(row=r, column=column_index_from_string('AI'), value=None)

            # Предрасчёт сессий: сгруппируем по пользователю и дате
            from collections import defaultdict
            sessions_by_user_date = defaultdict(lambda: defaultdict(float))
            for s in qs.select_related('user'):
                d = s.start_time.date()
                sessions_by_user_date[s.user_id][d] += s.duration_seconds() / 3600.0

            # Ручные отметки по пользователям
            marks_qs = WorkDayMark.objects.filter(date__gte=df, date__lte=dt)
            marks_map = defaultdict(dict)
            for m in marks_qs:
                marks_map[m.user_id][m.date] = m.code

            # Итерация по пользователям
            row_idx = start_row
            for u in users:
                full_name = (getattr(u, 'full_name', None) or u.get_full_name() or u.username)
                position_name = getattr(getattr(u, 'position', None), 'name', '') or ''
                ws.cell(row=row_idx, column=column_index_from_string('B'), value=full_name)
                ws.cell(row=row_idx, column=column_index_from_string('C'), value=position_name)

                # Заполним дни 1..31 в пределах выбранного периода
                # Колонки соответствуют дням МЕСЯЦА target_month (D = 1, ..., AH = 31)
                # Проходим по всем датам в периоде и маппим их на колонки
                from datetime import timedelta
                current_date = df
                while current_date <= dt:
                    # Определяем день месяца для маппинга на колонку
                    if current_date.year == target_year and current_date.month == target_month:
                        # Дата в целевом месяце - маппим напрямую
                        day_num = current_date.day
                        col = start_col + (day_num - 1)
                    else:
                        # Дата вне целевого месяца - пропускаем (не попадает в шаблон)
                        current_date += timedelta(days=1)
                        continue
                    
                    if day_num > days_in_month:
                        # Дня нет в целевом месяце
                        current_date += timedelta(days=1)
                        continue
                    
                    day_date = current_date

                    # Если есть ручная отметка — ставим её и пропускаем расчёт часов
                    mark = marks_map.get(u.id, {}).get(day_date)
                    if mark:
                        ws.cell(row=row_idx, column=col, value=mark)
                        current_date += timedelta(days=1)
                        continue

                    hours = sessions_by_user_date.get(u.id, {}).get(day_date, 0.0)
                    if hours > 0:
                        # Точные часы без округления
                        ws.cell(row=row_idx, column=col, value=round(hours, 2))
                    # Если нет часов и нет ручной отметки — оставляем пусто
                    
                    current_date += timedelta(days=1)

                # Формула итога по явочным часам
                ws['AI{}'.format(row_idx)] = '=SUM(D{r}:AH{r})'.format(r=row_idx)
                row_idx += 1

        else:
            # Fallback: простая книга
            logger.warning(f'Template not found at {template_path}, using fallback')
            wb = Workbook()
            ws = wb.active
            ws.title = "Timeclock"
            headers = ['Дата', 'Пользователь', 'Начало', 'Конец', 'Часы', 'Примечание']
            ws.append(headers)
            for s in qs.select_related('user').order_by('user__id', 'start_time'):
                local_start = make_naive(s.start_time)
                local_end = make_naive(s.end_time)
                hours = round(s.duration_seconds() / 3600, 2)
                ws.append([
                    local_start.date(),
                    s.user.get_full_name() or s.user.username,
                    local_start,
                    local_end,
                    hours,
                    s.note or '',
                ])

            if ws.max_row > 1:
                sum_row = ws.max_row + 1
                ws.cell(row=sum_row, column=5, value=f"=SUM(E2:E{ws.max_row})")

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        filename = f"Табель_{date_from}_to_{date_to}.xlsx"
        response = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        # В случае любой ошибки рендерим минимальный экспорт, чтобы не падать
        logger.error(f'Error during export: {e}', exc_info=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Timeclock"
        headers = ['Дата', 'Пользователь', 'Начало', 'Конец', 'Часы', 'Примечание']
        ws.append(headers)
        for s in qs.select_related('user').order_by('user__id', 'start_time'):
            local_start = timezone.localtime(s.start_time, timezone.get_current_timezone()).replace(tzinfo=None)
            local_end = (
                timezone.localtime(s.end_time, timezone.get_current_timezone()).replace(tzinfo=None)
                if s.end_time else None
            )
            hours = round(s.duration_seconds() / 3600, 2)
            ws.append([
                local_start.date(),
                s.user.get_full_name() or s.user.username,
                local_start,
                local_end,
                hours,
                s.note or '',
            ])
        if ws.max_row > 1:
            sum_row = ws.max_row + 1
            ws.cell(row=sum_row, column=5, value=f"=SUM(E2:E{ws.max_row})")
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        filename = f"Timeclock_{date_from}_to_{date_to}.xlsx"
        response = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def current_session_status(request):
    """Получить статус текущей открытой сессии пользователя."""
    user = request.user
    session = WorkSession.objects.filter(user=user, is_closed=False).first()
    if not session:
        return Response({'has_session': False})
    return Response({
        'has_session': True,
        'session_id': session.id,
        'start_time': session.start_time,
        'last_activity': session.last_activity,
        'duration_hours': session.duration_hours(),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_sessions(request):
    """Список сессий текущего пользователя."""
    user = request.user
    limit = int(request.GET.get('limit', 30))
    sessions = WorkSession.objects.filter(user=user).order_by('-start_time')[:limit]
    data = []
    for s in sessions:
        data.append({
            'id': s.id,
            'start_time': s.start_time,
            'end_time': s.end_time,
            'is_closed': s.is_closed,
            'duration_hours': s.duration_hours(),
            'note': s.note,
        })
    return Response({'sessions': data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_marks(request):
    """Получить ручные отметки пользователя за период."""
    date_from = request.GET.get('from')
    date_to = request.GET.get('to')
    if not date_from or not date_to:
        return Response({'detail': 'Missing dates'}, status=400)
    from datetime import datetime
    # Поддержка ISO 8601 с временем/таймзоной (берём первые 10 символов YYYY-MM-DD)
    df = datetime.strptime((date_from or '')[:10], '%Y-%m-%d').date()
    dt = datetime.strptime((date_to or '')[:10], '%Y-%m-%d').date()
    uid = int(request.GET.get('user_id') or request.user.id)
    marks = WorkDayMark.objects.filter(user_id=uid, date__gte=df, date__lte=dt)
    return Response({'marks': [{
        'date': m.date.isoformat(),
        'code': m.code
    } for m in marks]})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def set_mark(request):
    """Установить/обновить отметку дня (К/Б/А/О/В)."""
    from datetime import datetime
    data = request.data if hasattr(request, 'data') else request.POST
    date_str = data.get('date')
    code = data.get('code') or ''
    # Принимаем кириллицу напрямую, или конвертируем латиницу для обратной совместимости
    code_map = {'K': 'К', 'B': 'Б', 'A': 'А', 'O': 'О', 'V': 'В'}
    if code in code_map:
        code = code_map[code]
    uid = int(data.get('user_id') or request.user.id)
    if code not in ('К', 'Б', 'А', 'О', 'В'):
        return Response({'detail': 'Invalid code'}, status=400)
    try:
        d = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        return Response({'detail': 'Invalid date'}, status=400)
    mark, _ = WorkDayMark.objects.update_or_create(user_id=uid, date=d, defaults={'code': code})
    return Response({'status': 'ok', 'date': mark.date, 'code': mark.code})

